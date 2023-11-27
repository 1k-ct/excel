import openpyxl

def getList():
    excel_file_path = './test.xlsx'
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    # シートを選択
    sheet = workbook["Sheet1"]  # または workbook['Sheet1'] のようにシート名を指定

    # A列とB列のデータをセットで取得
    data_set = [(cell_A.value, cell_B.value, cell_D.value) for cell_A, cell_B, cell_D in zip(sheet['A'], sheet['C'],sheet['D'])]

  
    print(data_set)
    
    summary_dict = {}
    for item in data_set:
        key = (item[0], item[1])
        value = item[2]
        if key not in summary_dict:
            summary_dict[key] = [value]
        else:
            summary_dict[key].append(value)

    # 結果を表示
    for key, value in summary_dict.items():
        print(key, value)
    print(summary_dict)

    sheet2 = workbook['Sheet2']
    
    data = makeData(summary_dict)
    
    activeSheet = workbook['Sheet2']
    for i, (key, values) in enumerate(data.items(), start=1):
        activeSheet.cell(row=i, column=1).value = key[0]
        activeSheet.cell(row=i, column=2).value = key[1]
        for j, value in enumerate(values, start=3):
            activeSheet.cell(row=i, column=j).value = value
        workbook.save(excel_file_path)
        workbook.close()
    

def insertSheet(data):
    excel_file_path = './practice-vba.xlsm'
    workbook = openpyxl.load_workbook(excel_file_path)
    activeSheet = workbook['Sheet2']
    for i, (key, values) in enumerate(data.items(), start=1):
        activeSheet.cell(row=i, column=1).value = key[0]
        activeSheet.cell(row=i, column=2).value = key[1]
        for j, value in enumerate(values, start=3):
            activeSheet.cell(row=i, column=j).value = value
    workbook.save(excel_file_path)
    workbook.close()
# data_set = getList()
# insertSheet(sheet="Sheet2", data=data_set)

def makeData(summary_dict_1):
    new_summary_dict_1 = {}

    for key, values in summary_dict_1.items():
        chunked_values = [values[i:i + 20] for i in range(0, len(values), 20)]
        for i, chunk in enumerate(chunked_values):
            new_key = key if i == 0 else (key[0] + f"_{i}", key[1])
            new_summary_dict_1[new_key] = chunk
    return new_summary_dict_1

getList()