import openpyxl

wb = openpyxl.load_workbook("./practice-vba.xlsm")

sheet = wb["Sheet1"]

def getRowIndex():
    """_summary_
        A1からA7
    Args:
        row (_type_): _description_

    Returns:
        _type_: _description_
    """    
    index = []
    for col in sheet.iter_rows(min_row=2,min_col=1):
        index.append(col[0].value)
    
    return index

def getRowData(row):
    """_summary_
        
    Args:
        row (_type_): _description_

    Returns:
        _type_: _description_
    """    
    index = []
    for col in sheet.iter_rows(min_row=2,min_col=2):
        index.append(col[row].value)
    
    return index

def getIndexColumns():
    """_summary_
        BからF
    Args:
        column (_type_): _description_

    Returns:
        _type_: _description_
    """    
    index = []
    for col in sheet.iter_cols(min_col=2,max_col=6,min_row=1,max_row=7):
        index.append(col[0].value)
    
    return index

def main():
    tmp = []
    # for i in range(len(getIndexColumns())):
    # print("rowData",rowData)
    rowIndex = getRowIndex()
    # print("rowIndex",rowIndex)
    indexColumns = getIndexColumns()
    # print("indexColumns",indexColumns)
    for i in range(len(indexColumns)):
        rowData = getRowData(i)
        for j in range(len(rowData)):
            data = []
            if rowData[j] != None:
                data.append(indexColumns[i])
                data.append(rowIndex[j])
                data.append(rowData[j])
                tmp.append(data)
    for i in tmp:
        print(i)
sheet_range = sheet['B1':'F1']
# print(sheet_range[0][1].value)
main()
wb.close()